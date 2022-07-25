from csv import excel
from bs4 import BeautifulSoup
import requests
import openpyxl

excel=openpyxl.Workbook()
print(excel.sheetnames)
sheet =excel.active
sheet.title='IMDB TOP 250 MOVIES'
print(excel.sheetnames)
sheet.append(['Movie Rank','Movie Name','Movie year','IMDB rating'])

try:
    source =requests.get('https://www.imdb.com/chart/top/?ref_=nv_mv_250')
    source.raise_for_status()

    soup=BeautifulSoup(source.text,'html.parser')
    movies = soup.find('tbody',class_="lister-list").find_all('tr')
    for movie in movies:
        Name= movie.find('td',class_="titleColumn").a.text
        Rank= movie.find('td',class_="titleColumn").get_text(strip=True).split('.')[0]
        Year = movie.find('td',class_="titleColumn").span.text.strip('()')
        Rating = movie.find('td',class_="ratingColumn imdbRating").strong.text
        print(Rank, Name, Year, Rating)
        sheet.append([Rank, Name, Year, Rating])


except Exception as e:
    print(e)

excel.save('IMDB movies Rating.xlsx')
